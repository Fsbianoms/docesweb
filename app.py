from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import pandas as pd
import openpyxl
import os
import datetime
import math

app = Flask(__name__, static_folder=os.getcwd(), static_url_path='')
CORS(app)

EXCEL_FILE = "Estou compartilhando o arquivo 'Janeiro-2026 Atual-3' com você.xlsx"

def is_nan(value):
    if isinstance(value, float) and math.isnan(value):
        return True
    return False

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/<path:path>')
def static_files(path):
    return send_from_directory('.', path)

@app.route('/api/sync', methods=['POST'])
def sync_data():
    try:
        import subprocess
        data = request.json or {}
        sheet_url = data.get('url')
        
        import sys
        python_exe = sys.executable
            
        cmd_args = [python_exe, "parse_data.py"]
        if sheet_url:
            cmd_args.append(sheet_url)
            
        result = subprocess.run(cmd_args, capture_output=True, text=True)
        
        if result.returncode == 0:
            return jsonify({"success": True, "message": "Dados atualizados com sucesso!"})
        else:
            print(f"Erro no parse_data: {result.stderr}")
            return jsonify({
                "success": False, 
                "message": "Erro ao processar planilha.",
                "details": result.stderr
            })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/venda', methods=['POST'])
def add_sale():
    try:
        data = request.json
        # ... (rest of logic same as before)
        
        # ... logic to get month, sheet_name etc ...
        dt = datetime.datetime.strptime(data.get('data'), '%Y-%m-%d')
        month_names = {1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto", 9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"}
        sheet_name = month_names[dt.month]

        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            if sheet_name not in wb.sheetnames:
                return jsonify({"success": False, "message": f"Aba '{sheet_name}' não encontrada."}), 400
            
            sheet = wb[sheet_name]
            # ... encontrar cabecalho e colunas ...
            header_row = -1
            col_map = {}
            for r in range(1, 10):
                row_values = [str(sheet.cell(row=r, column=c).value).strip() if sheet.cell(row=r, column=c).value else "" for c in range(1, sheet.max_column + 1)]
                if 'Cliente' in row_values:
                    header_row = r
                    for c in range(1, sheet.max_column + 1):
                        val_lower = str(sheet.cell(row=r, column=c).value).strip().lower() if sheet.cell(row=r, column=c).value else ""
                        if 'cliente' in val_lower: col_map['cliente'] = c
                        elif 'produto' in val_lower: col_map['produto'] = c
                        elif 'quant' in val_lower: col_map['quantidade'] = c
                        elif 'valor' in val_lower: col_map['valor'] = c
                        elif 'status' in val_lower: col_map['status'] = c
                        elif 'data da venda' in val_lower: col_map['data_venda'] = c
                        elif 'contato' in val_lower: col_map['contato'] = c
                        elif 'observ' in val_lower and 'contato' not in col_map: col_map['contato'] = c
                    break
            
            next_row = header_row + 1
            while sheet.cell(row=next_row, column=col_map['cliente']).value is not None:
                next_row += 1
            
            sheet.cell(row=next_row, column=col_map['cliente']).value = data.get('cliente')
            if 'contato' in col_map:
                sheet.cell(row=next_row, column=col_map['contato']).value = data.get('contato')
            sheet.cell(row=next_row, column=col_map['produto']).value = data.get('produto')
            sheet.cell(row=next_row, column=col_map['quantidade']).value = data.get('quantidade')
            sheet.cell(row=next_row, column=col_map['valor']).value = float(data.get('valor'))
            sheet.cell(row=next_row, column=col_map['status']).value = data.get('status').upper()
            sheet.cell(row=next_row, column=col_map['data_venda']).value = dt
            
            wb.save(EXCEL_FILE)
            
            python_exe = os.path.join('.venv', 'Scripts', 'python.exe')
            if not os.path.exists(python_exe):
                python_exe = 'python'
                
            os.system(f'"{python_exe}" parse_data.py')
            return jsonify({"success": True, "message": "Venda registrada com sucesso!"})
            
        except PermissionError:
            return jsonify({"success": False, "message": "ACESSO NEGADO: Feche o arquivo Excel antes de registrar a venda!"}), 403

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 8080))
    print(f"Iniciando servidor API DocesWeb na porta {port}...")
    app.run(host='0.0.0.0', port=port)
